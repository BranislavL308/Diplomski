from PyQt5.QtWidgets import QMainWindow, QFileDialog, QMessageBox, QGridLayout, QTableWidget, QTableWidgetItem, QVBoxLayout, QApplication, QCheckBox, QLabel, QInputDialog, QLineEdit
from mainWindowUI import *
from settingsDialog import *
from communicationThread import *
from PyQt5 import QtWidgets
from PyQt5.QtCore import QPoint
import PyQt5.QtCore as QtCore
from PyQt5 import QtGui
from PyQt5.QtCore import pyqtSignal 
from calDialog import *
import time
import os,sys
from measure_thread import * 
import pandas as pd 
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from datetime import datetime
from PyQt5.QtGui import QIcon, QPixmap



class MyForm(QMainWindow):
    send_data = pyqtSignal(dict)
    send_izmereno = pyqtSignal(float)
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        #self.com1 = com1
        self.returnVal = {}
        
        
        self.ui.pushButtonStart.setEnabled(False)
        self.ui.pushButtonStop.setEnabled(False)

        self.ui.pushButtonPodesavanja.clicked.connect(self.config)
        self.ui.pushButtonStart.clicked.connect(self.open_com)
        self.ui.pushButtonStop.clicked.connect(self.close_com)

        self.podaciMain ={}

        self.com_thread = MyThread(self.returnVal)
        self.measure_thread = MeasureThread()
        
        self.ui.pushButtonSet.clicked.connect(self.measure_thread.update_flag)
        self.ui.pushButtonSet.clicked.connect(self.getMessure)
        self.measure_thread.set_data.connect(self.set_data)
        self.measure_thread.send_measure.connect(self.getData)
        
        self.send_data.connect(self.measure_thread.collect_data)
        self.send_izmereno.connect(self.measure_thread.collect_measure)
        
        self.ui.pushButtonSave.setEnabled(False)
        self.ui.pushButtonSave.clicked.connect(self.save_data)
        
        self.ui.pushButtonSet.setEnabled(False)

        self.flag = False
         



    def set_data(self, str1, str2, str3, str4):
        self.ui.lineEditTip.setText(str1)
        self.ui.lineEditRng.setText(str2)
        try:
            self.currentRng = round(float(str2),1)
        except:
            pass

        self.ui.lineEditVal.setText(str3)
        self.ui.lineEditFreq.setText(str4)

    def calCheck(self):
        self.calDialog = CalDialog(self.measure_thread.dataType,self.currentRng)
        self.calDialog.exec_()

    def printData(self):
        return self.podaciMain

    def config(self):
        self.ui.pushButtonStart.setEnabled(False)
        self.ui.pushButtonStop.setEnabled(False)
        
        self.settingsDialog = SettingsDialog()
        
        response = self.settingsDialog.exec_() #funkcija .exec_ pokrece nas prozor
        #Ako pritisnemo x dobijemo 0
        #Ako pritisnemo OK dobijemo 1

        #Ako je korisnik odgovorio sa response onda pozivamo metodu
        if(response):
            self.returnVal = self.settingsDialog.get_data()
            self.ui.pushButtonStart.setEnabled(True)
            self.ui.pushButtonStop.setEnabled(False)
            
            self.podaciMain = self.settingsDialog.checkData()
            self.ui.pushButtonSave.setEnabled(False)

        else:
            self.returnVal['com_port'] = ""

    
    def open_com(self):
        self.ui.pushButtonStart.setEnabled(False)
        self.ui.pushButtonStop.setEnabled(True)
        self.com_thread.init_com(self.returnVal)
        self.com_thread.error.connect(self.error)
        self.send_data.emit(self.podaciMain)
        self.com_thread.start()
        self.measure_thread.start()
        self.measure_thread.send_data.connect(self.com_thread.sendData)
        self.ui.pushButtonSet.setEnabled(True)

    def getMessure(self):
        if(self.measure_thread.flag_end == False):
            self.calCheck()
            self.com_thread.sendData('OUTP ON\n\r')
            time.sleep(0.01)
            self.i, okPressed = QInputDialog.getDouble(self,"Izmerena vrednost","Value:", 0, -10000, 10000000000, 10)
            if(okPressed):
                self.send_izmereno.emit(self.i)
            self.com_thread.sendData('OUTP OFF\n\r')
            time.sleep(0.01)
            self.measure_thread.update_flag()
        else:
            self.clearTextLine()
            self.ui.pushButtonSave.setEnabled(True)
            self.close_com()
            self.measure_thread.flag_end = False
    
    def error(self):
        pass
     
    def getData(self, data):
        self.podaciMereno = data

    def clearTextLine(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setInformativeText('Kraj merenja.')
        msg.setWindowTitle('Kraj')
        msg.exec_()
        self.ui.pushButtonSet.setEnabled(False)
        self.ui.lineEditTip.setText('')
        self.ui.lineEditRng.setText('')
        self.ui.lineEditVal.setText('')
        self.ui.lineEditFreq.setText('')
        

    
    def close_com(self):
        self.ui.pushButtonStart.setEnabled(True)
        self.ui.pushButtonStop.setEnabled(False)
        self.com_thread.exit()
        self.measure_thread.exit()
         



    def save_data(self):
          
        self.dataRES = self.podaciMain.get("RES")
        self.dataACV = self.podaciMain.get("ACV")
        self.dataACI = self.podaciMain.get("ACI")
        self.dataDCV = self.podaciMain.get("DCV")
        self.dataDCI = self.podaciMain.get("DCI")
        self.dataFREQ = self.podaciMain.get("FREQ")
        self.dataTEMP = self.podaciMain.get("TEMP")
        #Mereni podaci
        self.dataRES_M = self.podaciMereno.get("RES")
        self.dataACV_M = self.podaciMereno.get("ACV")
        self.dataACI_M = self.podaciMereno.get("ACI")
        self.dataDCV_M = self.podaciMereno.get("DCV")
        self.dataDCI_M = self.podaciMereno.get("DCI")
        self.dataFREQ_M = self.podaciMereno.get("FREQ")
        self.dataTEMP_M = self.podaciMereno.get("TEMP")
        
        #filename = QFileDialog.getSaveFileName()

        self.cnt_sheet = 0


     

        book = Workbook()
        sheet = book.active

        #Definisanje linija
        thin = Side(style='thin')
        thick = Side(style='thick')
        double = Side(style='double')
        medium = Side(style='medium')
        
      
        #DC napon
        if(len(self.dataDCV) != 0):
            sheet = book.create_sheet("Multimetar DCV", self.cnt_sheet)
            self.cnt_sheet +=1

            sheet['A3'].border = Border(top=thin,bottom=thin,right=thin,left=thin) 
            sheet['B3'].border = Border(top=thin,bottom=thin,right=thin,left=thin) 
            sheet['C3'].border = Border(top=thin,bottom=thin,right=None,left=thin) 
            sheet['C3'].border = Border(top=thin,bottom=thin,right=None,left=None)
            sheet['D3'].border = Border(top=thin,bottom=thin,right=thin,left=thin)
            sheet['E3'].border = Border(top=thin,bottom=thin,right=None,left=thin) 
            sheet['E3'].border = Border(top=thin,bottom=thin,right=None,left=None)
            sheet['E3'].border = Border(top=thin,bottom=thin,right=None,left=None)  

            sheet['A4'].border = Border(right=thin) 
            sheet['B4'].border = Border(right=thin) 
            sheet['C4'].border = Border(right=thin) 
            sheet['D4'].border = Border(right=thin)
            sheet['E4'].border = Border(right=thin)  

            sheet['A5'].border = Border(bottom=double,right=thin) 
            sheet['B5'].border = Border(bottom=double,right=thin) 
            sheet['C5'].border = Border(bottom=double,right=thin) 
            sheet['D5'].border = Border(bottom=double,right=thin)
            sheet['E5'].border = Border(bottom=double,right=thin)  

            sheet['A6'].border = Border(right=thin, top = None) 
            sheet['B6'].border = Border(right=thin) 
            sheet['C6'].border = Border(right=thin, top = None) 
            sheet['D6'].border = Border(right=thin)
            sheet['E6'].border = Border(right=thin) 

            
            sheet['A7'].border = Border(right=thin) 
            sheet['B7'].border = Border(right=thin) 
            sheet['C7'].border = Border(right=thin) 
            sheet['D7'].border = Border(right=thin)
            sheet['E7'].border = Border(right=thin) 

            sheet['A8'].border = Border(right=thin) 
            sheet['B8'].border = Border(right=thin) 
            sheet['C8'].border = Border(right=thin) 
            sheet['D8'].border = Border(right=thin)
            sheet['E8'].border = Border(right=thin) 

            #sirina celova
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 20
            sheet.column_dimensions['D'].width = 20
            sheet.column_dimensions['E'].width = 20

            #Visina 
            sheet.row_dimensions[4].height = 30
        
            
            cell = sheet['A2']
            cell.value = 'Jednosmerni Napon'


            cell = sheet['A3']
            cell.value = 'Etalon'
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            
            sheet.merge_cells('B3:C3')
            cell = sheet['B3']
            cell.value = 'OE'
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            sheet.merge_cells('D3:E3')
            cell = sheet['D3']
            cell.value = 'Obrada rezultata'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet['A4']
            cell.value = 'Pokazivanje etalona'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet['B4']
            cell.value = 'Opseg instrumenta'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet['C4']
            cell.value = 'Očitana vrednost sa instrumenta'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)


            cell = sheet['D4']
            cell.value = 'G'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet['E4']
            cell.value = 'U'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet['A5']
            cell.value = 'V' #Jedinice merene velicine
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet['B5']
            cell.value = 'V'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet['C5']
            cell.value = 'V'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet['D5']
            cell.value = 'V'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet['B5']
            cell.value = 'V'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet['E5']
            cell.value = 'V'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')


            cell = sheet['A34']
            cell.value = 'OE'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet['A35']
            cell.value = 'G'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet['A36']
            cell.value = 'U'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet['C34']
            cell.value = 'Objekt etaloniranja'

            cell = sheet['C35']
            cell.value = 'Greška OE'

            cell = sheet['C36']
            cell.value = 'Proširena (k=2) merna nesigurnost etaloniranja'


            sheet.merge_cells('A39:D39')
            cell = sheet['A39']
            cell.value = 'Merna nesigurnost iskazana u ovom Uverenju je proširena merna nesigurnost, gde je standardna'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet.merge_cells('A40:D40')
            cell = sheet['A40']
            cell.value = 'merna nesigurnost pomnožena faktorom obuhvata k = 2, što za slučaj normalne raspodele greške'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet.merge_cells('A41:D41')
            cell = sheet['A41']
            cell.value = 'odgovara verovatnoći od približno 95 %.'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet.merge_cells('A43:D43')
            cell = sheet['A43']
            cell.value = '- Kraj Uverenja -'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(size=10)

            #********************************************************
            #Kraj jedne stranice
            #Ubacivanje podataka u stranicu
            if(len(self.dataDCV_M) != 0):
                a = len(self.dataDCV)
                flag = True
                tmpDCV = []
                tmpDCV_M = []
                cnt = 4
                for i in range(a):
                    tmpData = self.dataDCV[i][0]
                    tmpData_M = self.dataDCV_M[i][0]
                    if(flag == True):
                        flag = False
                        sheet.cell(row=6, column=2).value = self.dataDCV[i][1]
                    elif(len(tmpData) == 5):
                        sheet.cell(row=6+cnt, column=2).value = self.dataDCV[i][1]
                        cnt +=5
                    else:
                        sheet.cell(row=6+cnt, column=2).value = self.dataDCV[i][1]
                        cnt +=4 

                    r = len(tmpData)
                    for j in range(r):
                        tmpDCV.append(tmpData[j])
                        tmpDCV_M.append(tmpData_M[j])
        
            else:
                pass

            for i in range(len(tmpDCV)):
                sheet.cell(row=6+i, column=1).value = tmpDCV[i]
                sheet.cell(row=6+i, column=3).value = tmpDCV_M[i]

                #Greska
                sheet.cell(row=6+i, column=4).value = tmpDCV_M[i] - tmpDCV[i]

                sheet.cell(row=6+i, column=1).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet.cell(row=6+i, column=2).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet.cell(row=6+i, column=3).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet.cell(row=6+i, column=4).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet.cell(row=6+i, column=5).border = Border(top=None,bottom=None,right=thin,left=None)

                sheet.cell(row=6+i, column=1).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet.cell(row=6+i, column=2).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet.cell(row=6+i, column=3).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet.cell(row=6+i, column=4).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet.cell(row=6+i, column=5).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        else:
            pass   
        
        #AC napon    
        if(len(self.dataACV) != 0):
            sheet3 = book.create_sheet("Multimetar ACV", self.cnt_sheet)
            self.cnt_sheet +=1

            sheet3['A3'].border = Border(top=thin,bottom=thin,right=thin,left=thin) 
            sheet3['B3'].border = Border(top=thin,bottom=thin,right=thin,left=thin) 
            sheet3['C3'].border = Border(top=thin,bottom=thin,right=None,left=thin) 
            sheet3['C3'].border = Border(top=thin,bottom=thin,right=None,left=None)
            sheet3['D3'].border = Border(top=thin,bottom=thin,right=thin,left=thin)
            sheet3['E3'].border = Border(top=thin,bottom=thin,right=None,left=thin) 
            sheet3['E3'].border = Border(top=thin,bottom=thin,right=None,left=None)
            sheet3['E3'].border = Border(top=thin,bottom=thin,right=None,left=None)  
            sheet3['F3'].border = Border(top=None,bottom=None,right=thin,left=None)

            sheet3['A4'].border = Border(right=thin) 
            sheet3['B4'].border = Border(right=thin) 
            sheet3['C4'].border = Border(right=thin) 
            sheet3['D4'].border = Border(right=thin)
            sheet3['E4'].border = Border(right=thin)
            sheet3['F4'].border = Border(right=thin) 

            sheet3['A5'].border = Border(bottom=double,right=thin) 
            sheet3['B5'].border = Border(bottom=double,right=thin) 
            sheet3['C5'].border = Border(bottom=double,right=thin) 
            sheet3['D5'].border = Border(bottom=double,right=thin)
            sheet3['E5'].border = Border(bottom=double,right=thin)  
            sheet3['F5'].border = Border(bottom=double,right=thin)

            sheet3['A6'].border = Border(right=thin, top = None) 
            sheet3['B6'].border = Border(right=thin) 
            sheet3['C6'].border = Border(right=thin, top = None) 
            sheet3['D6'].border = Border(right=thin)
            sheet3['E6'].border = Border(right=thin) 
            sheet3['F6'].border = Border(right=thin)
            
            sheet3['A7'].border = Border(right=thin) 
            sheet3['B7'].border = Border(right=thin) 
            sheet3['C7'].border = Border(right=thin) 
            sheet3['D7'].border = Border(right=thin)
            sheet3['E7'].border = Border(right=thin) 
            sheet3['F7'].border = Border(right=thin)

            sheet3['A8'].border = Border(right=thin) 
            sheet3['B8'].border = Border(right=thin) 
            sheet3['C8'].border = Border(right=thin) 
            sheet3['D8'].border = Border(right=thin)
            sheet3['E8'].border = Border(right=thin) 
            sheet3['F8'].border = Border(right=thin)
            #sirina celova
            sheet3.column_dimensions['A'].width = 20
            sheet3.column_dimensions['B'].width = 20
            sheet3.column_dimensions['C'].width = 20
            sheet3.column_dimensions['D'].width = 20
            sheet3.column_dimensions['E'].width = 20
            sheet3.column_dimensions['F'].width = 20

            #Visina 
            sheet3.row_dimensions[4].height = 30
        
            
            cell = sheet3['A2']
            cell.value = 'Naizmenični napon'


            sheet3.merge_cells('A3:B3')
            cell = sheet3['A3']
            cell.value = 'Etalon'
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            
            sheet3.merge_cells('C3:D3')
            cell = sheet3['C3']
            cell.value = 'OE'
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            sheet3.merge_cells('E3:F3')
            cell = sheet3['E3']
            cell.value = 'Obrada rezultata'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet3['A4']
            cell.value = 'Frekvencija'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet3['B4']
            cell.value = 'Pokazivanje etalona'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet3['C4']
            cell.value = 'Opseg instrumenta'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)


            cell = sheet3['D4']
            cell.value = 'Očitana vrednost sa instrumenta'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet3['E4']
            cell.value = 'G'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            cell = sheet3['F4']
            cell.value = 'U'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet3['A5']
            cell.value = 'Hz' #Jedinice merene velicine
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet3['B5']
            cell.value = 'V'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet3['C5']
            cell.value = 'V'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet3['D5']
            cell.value = 'V'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet3['B5']
            cell.value = 'V'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet3['E5']
            cell.value = 'V'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            
            cell = sheet3['F5']
            cell.value = 'V'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet3['A34']
            cell.value = 'OE'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet3['A35']
            cell.value = 'G'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet3['A36']
            cell.value = 'U'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet3['C34']
            cell.value = 'Objekt etaloniranja'

            cell = sheet3['C35']
            cell.value = 'Greška OE'

            cell = sheet3['C36']
            cell.value = 'Proširena (k=2) merna nesigurnost etaloniranja'


            sheet3.merge_cells('A39:D39')
            cell = sheet3['A39']
            cell.value = 'Merna nesigurnost iskazana u ovom Uverenju je proširena merna nesigurnost, gde je standardna'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet3.merge_cells('A40:D40')
            cell = sheet3['A40']
            cell.value = 'merna nesigurnost pomnožena faktorom obuhvata k = 2, što za slučaj normalne raspodele greške'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet3.merge_cells('A41:D41')
            cell = sheet3['A41']
            cell.value = 'odgovara verovatnoći od približno 95 %.'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet3.merge_cells('A43:D43')
            cell = sheet3['A43']
            cell.value = '- Kraj Uverenja -'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(size=10)

            #********************************************************
            #Kraj jedne stranice
            #Ubacivanje podataka u stranicu 
            if(len(self.dataACV_M) != 0):
                a = len(self.dataACV)
                flag1 = True
                tmpACV = []
                tmpACV_M = []
                tmpACV_F = []
                cnt = 0
                for i in range(a):
                    tmpData = self.dataACV[i][0] #Vrednost
                    tmpData_M = self.dataACV_M[i][0] #Vrednost merena
                    #tmpData_F = self.dataACV[i][1]
                    if(flag1 == True):
                        flag1 = False
                        sheet3.cell(row=6, column=3).value = self.dataACV[i][2]
                        sheet3.cell(row=6, column=1).value = self.dataACV[i][1]
                        sheet3.cell(row=7, column=1).value = self.dataACV[i][1]
                        cnt +=2
                    elif(len(tmpData) == 3):
                        sheet3.cell(row=6+cnt, column=3).value = self.dataACV[i][2]
                        sheet3.cell(row=6+cnt, column=1).value = self.dataACV[i][1]
                        sheet3.cell(row=7+cnt, column=1).value = self.dataACV[i][1]
                        sheet3.cell(row=8+cnt, column=1).value = self.dataACV[i][1]
                        cnt +=3
                    else:
                        sheet3.cell(row=6+cnt, column=3).value = self.dataACV[i][2]
                        sheet3.cell(row=6+cnt, column=1).value = self.dataACV[i][1]
                        cnt +=1


                    #tmpACV_F.append(tmpData_F) 
                    r = len(tmpData)
                    for j in range(r):
                        tmpACV.append(tmpData[j])
                        tmpACV_M.append(tmpData_M[j])
                        
                    #sheet3.cell(row=6+i, column=2).value = tmpDCV[i]
            else:
                pass

            for i in range(len(tmpACV)):
                sheet3.cell(row=6+i, column=2).value = tmpACV[i]
                sheet3.cell(row=6+i, column=4).value = tmpACV_M[i]

                #Greska
                sheet3.cell(row=6+i, column=5).value = tmpACV_M[i] - tmpACV[i]

                sheet3.cell(row=6+i, column=1).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet3.cell(row=6+i, column=2).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet3.cell(row=6+i, column=3).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet3.cell(row=6+i, column=4).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet3.cell(row=6+i, column=5).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet3.cell(row=6+i, column=6).border = Border(top=None,bottom=None,right=thin,left=None)

                sheet3.cell(row=6+i, column=1).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet3.cell(row=6+i, column=2).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet3.cell(row=6+i, column=3).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet3.cell(row=6+i, column=4).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet3.cell(row=6+i, column=5).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet3.cell(row=6+i, column=6).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        else:
            pass

        #DC struja
        if(len(self.dataDCI) != 0):
            
            sheet1 = book.create_sheet("Multimetar DCI", self.cnt_sheet)
            self.cnt_sheet +=1

            sheet1['A3'].border = Border(top=thin,bottom=thin,right=thin,left=thin) 
            sheet1['B3'].border = Border(top=thin,bottom=thin,right=thin,left=thin) 
            sheet1['C3'].border = Border(top=thin,bottom=thin,right=None,left=thin) 
            sheet1['C3'].border = Border(top=thin,bottom=thin,right=None,left=None)
            sheet1['D3'].border = Border(top=thin,bottom=thin,right=thin,left=thin)
            sheet1['E3'].border = Border(top=thin,bottom=thin,right=None,left=thin) 
            sheet1['E3'].border = Border(top=thin,bottom=thin,right=None,left=None)
            sheet1['E3'].border = Border(top=thin,bottom=thin,right=None,left=None)  

            sheet1['A4'].border = Border(right=thin) 
            sheet1['B4'].border = Border(right=thin) 
            sheet1['C4'].border = Border(right=thin) 
            sheet1['D4'].border = Border(right=thin)
            sheet1['E4'].border = Border(right=thin)  

            sheet1['A5'].border = Border(bottom=double,right=thin) 
            sheet1['B5'].border = Border(bottom=double,right=thin) 
            sheet1['C5'].border = Border(bottom=double,right=thin) 
            sheet1['D5'].border = Border(bottom=double,right=thin)
            sheet1['E5'].border = Border(bottom=double,right=thin)  

            sheet1['A6'].border = Border(right=thin, top = None) 
            sheet1['B6'].border = Border(right=thin) 
            sheet1['C6'].border = Border(right=thin, top = None) 
            sheet1['D6'].border = Border(right=thin)
            sheet1['E6'].border = Border(right=thin) 

            
            sheet1['A7'].border = Border(right=thin) 
            sheet1['B7'].border = Border(right=thin) 
            sheet1['C7'].border = Border(right=thin) 
            sheet1['D7'].border = Border(right=thin)
            sheet1['E7'].border = Border(right=thin) 

            sheet1['A8'].border = Border(right=thin) 
            sheet1['B8'].border = Border(right=thin) 
            sheet1['C8'].border = Border(right=thin) 
            sheet1['D8'].border = Border(right=thin)
            sheet1['E8'].border = Border(right=thin) 

            #sirina celova
            sheet1.column_dimensions['A'].width = 20
            sheet1.column_dimensions['B'].width = 20
            sheet1.column_dimensions['C'].width = 20
            sheet1.column_dimensions['D'].width = 20
            sheet1.column_dimensions['E'].width = 20

            #Visina 
            sheet1.row_dimensions[4].height = 30
        
            
            cell = sheet1['A2']
            cell.value = 'Jednosmerna struja'


            cell = sheet1['A3']
            cell.value = 'Etalon'
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            
            sheet1.merge_cells('B3:C3')
            cell = sheet1['B3']
            cell.value = 'OE'
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            sheet1.merge_cells('D3:E3')
            cell = sheet1['D3']
            cell.value = 'Obrada rezultata'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet1['A4']
            cell.value = 'Pokazivanje etalona'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet1['B4']
            cell.value = 'Opseg instrumenta'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet1['C4']
            cell.value = 'Očitana vrednost sa instrumenta'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)


            cell = sheet1['D4']
            cell.value = 'G'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet1['E4']
            cell.value = 'U'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet1['A5']
            cell.value = 'A' #Jedinice merene velicine
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet1['B5']
            cell.value = 'A'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet1['C5']
            cell.value = 'A'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet1['D5']
            cell.value = 'A'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet1['B5']
            cell.value = 'A'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet1['E5']
            cell.value = 'A'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')


            cell = sheet1['A34']
            cell.value = 'OE'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet1['A35']
            cell.value = 'G'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet1['A36']
            cell.value = 'U'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet1['C34']
            cell.value = 'Objekt etaloniranja'

            cell = sheet1['C35']
            cell.value = 'Greška OE'

            cell = sheet1['C36']
            cell.value = 'Proširena (k=2) merna nesigurnost etaloniranja'


            sheet1.merge_cells('A39:D39')
            cell = sheet1['A39']
            cell.value = 'Merna nesigurnost iskazana u ovom Uverenju je proširena merna nesigurnost, gde je standardna'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet1.merge_cells('A40:D40')
            cell = sheet1['A40']
            cell.value = 'merna nesigurnost pomnožena faktorom obuhvata k = 2, što za slučaj normalne raspodele greške'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet1.merge_cells('A41:D41')
            cell = sheet1['A41']
            cell.value = 'odgovara verovatnoći od približno 95 %.'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet1.merge_cells('A43:D43')
            cell = sheet1['A43']
            cell.value = '- Kraj Uverenja -'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(size=10)

            #********************************************************
            #Kraj jedne stranice
            #Ubacivanje podataka u stranicu 
            if(len(self.dataDCI_M) != 0):
                a = len(self.dataDCI)
                flag1 = True
                tmpDCI = []
                tmpDCI_M = []
                cnt = 2
                for i in range(a):
                    tmpData = self.dataDCI[i][0]
                    tmpData_M = self.dataDCI_M[i][0]
                    if(flag1 == True):
                        flag1 = False
                        sheet1.cell(row=6, column=2).value = self.dataDCI[i][1]
                    else:
                        if(len(tmpData) == 4):
                            sheet1.cell(row=6+cnt, column=2).value = self.dataDCI[i][1]
                            cnt +=4 
                        else:
                            sheet1.cell(row=6+cnt, column=2).value = self.dataDCI[i][1]
                            cnt +=2 

                    r = len(tmpData)
                    for j in range(r):
                        tmpDCI.append(tmpData[j])
                        tmpDCI_M.append(tmpData_M[j])
                    #sheet1.cell(row=6+i, column=2).value = tmpDCV[i]
            else:
                pass

            for i in range(len(tmpDCI)):
                sheet1.cell(row=6+i, column=1).value = tmpDCI[i]
                sheet1.cell(row=6+i, column=3).value = tmpDCI_M[i]

                #Greska
                sheet1.cell(row=6+i, column=4).value = tmpDCI_M[i] - tmpDCI[i]

                sheet1.cell(row=6+i, column=1).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet1.cell(row=6+i, column=2).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet1.cell(row=6+i, column=3).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet1.cell(row=6+i, column=4).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet1.cell(row=6+i, column=5).border = Border(top=None,bottom=None,right=thin,left=None)

                sheet1.cell(row=6+i, column=1).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet1.cell(row=6+i, column=2).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet1.cell(row=6+i, column=3).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet1.cell(row=6+i, column=4).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet1.cell(row=6+i, column=5).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        else:
            pass

        #AC struja
        if(len(self.dataACI) != 0):
            sheet2 = book.create_sheet("Multimetar ACI", self.cnt_sheet)
            self.cnt_sheet +=1

            sheet2['A3'].border = Border(top=thin,bottom=thin,right=thin,left=thin) 
            sheet2['B3'].border = Border(top=thin,bottom=thin,right=thin,left=thin) 
            sheet2['C3'].border = Border(top=thin,bottom=thin,right=None,left=thin) 
            sheet2['C3'].border = Border(top=thin,bottom=thin,right=None,left=None)
            sheet2['D3'].border = Border(top=thin,bottom=thin,right=thin,left=thin)
            sheet2['E3'].border = Border(top=thin,bottom=thin,right=None,left=thin) 
            sheet2['E3'].border = Border(top=thin,bottom=thin,right=None,left=None)
            sheet2['E3'].border = Border(top=thin,bottom=thin,right=None,left=None)  
            sheet2['F3'].border = Border(top=None,bottom=None,right=thin,left=None)

            sheet2['A4'].border = Border(right=thin) 
            sheet2['B4'].border = Border(right=thin) 
            sheet2['C4'].border = Border(right=thin) 
            sheet2['D4'].border = Border(right=thin)
            sheet2['E4'].border = Border(right=thin)
            sheet2['F4'].border = Border(right=thin) 

            sheet2['A5'].border = Border(bottom=double,right=thin) 
            sheet2['B5'].border = Border(bottom=double,right=thin) 
            sheet2['C5'].border = Border(bottom=double,right=thin) 
            sheet2['D5'].border = Border(bottom=double,right=thin)
            sheet2['E5'].border = Border(bottom=double,right=thin)  
            sheet2['F5'].border = Border(bottom=double,right=thin)

            sheet2['A6'].border = Border(right=thin, top = None) 
            sheet2['B6'].border = Border(right=thin) 
            sheet2['C6'].border = Border(right=thin, top = None) 
            sheet2['D6'].border = Border(right=thin)
            sheet2['E6'].border = Border(right=thin) 
            sheet2['F6'].border = Border(right=thin)
            
            sheet2['A7'].border = Border(right=thin) 
            sheet2['B7'].border = Border(right=thin) 
            sheet2['C7'].border = Border(right=thin) 
            sheet2['D7'].border = Border(right=thin)
            sheet2['E7'].border = Border(right=thin) 
            sheet2['F7'].border = Border(right=thin)

            sheet2['A8'].border = Border(right=thin) 
            sheet2['B8'].border = Border(right=thin) 
            sheet2['C8'].border = Border(right=thin) 
            sheet2['D8'].border = Border(right=thin)
            sheet2['E8'].border = Border(right=thin) 
            sheet2['F8'].border = Border(right=thin)
            #sirina celova
            sheet2.column_dimensions['A'].width = 20
            sheet2.column_dimensions['B'].width = 20
            sheet2.column_dimensions['C'].width = 20
            sheet2.column_dimensions['D'].width = 20
            sheet2.column_dimensions['E'].width = 20
            sheet2.column_dimensions['F'].width = 20

            #Visina 
            sheet2.row_dimensions[4].height = 30
        
            
            cell = sheet2['A2']
            cell.value = 'Naizmenična struja'


            sheet2.merge_cells('A3:B3')
            cell = sheet2['A3']
            cell.value = 'Etalon'
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            
            sheet2.merge_cells('C3:D3')
            cell = sheet2['C3']
            cell.value = 'OE'
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            sheet2.merge_cells('E3:F3')
            cell = sheet2['E3']
            cell.value = 'Obrada rezultata'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet2['A4']
            cell.value = 'Frekvencija'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet2['B4']
            cell.value = 'Pokazivanje etalona'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet2['C4']
            cell.value = 'Opseg instrumenta'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)


            cell = sheet2['D4']
            cell.value = 'Očitana vrednost sa instrumenta'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet2['E4']
            cell.value = 'G'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            cell = sheet2['F4']
            cell.value = 'U'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet2['A5']
            cell.value = 'Hz' #Jedinice merene velicine
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet2['B5']
            cell.value = 'A'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet2['C5']
            cell.value = 'A'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet2['D5']
            cell.value = 'A'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet2['B5']
            cell.value = 'A'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet2['E5']
            cell.value = 'A'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            
            cell = sheet2['F5']
            cell.value = 'A'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet2['A34']
            cell.value = 'OE'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet2['A35']
            cell.value = 'G'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet2['A36']
            cell.value = 'U'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet2['C34']
            cell.value = 'Objekt etaloniranja'

            cell = sheet2['C35']
            cell.value = 'Greška OE'

            cell = sheet2['C36']
            cell.value = 'Proširena (k=2) merna nesigurnost etaloniranja'


            sheet2.merge_cells('A39:D39')
            cell = sheet2['A39']
            cell.value = 'Merna nesigurnost iskazana u ovom Uverenju je proširena merna nesigurnost, gde je standardna'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet2.merge_cells('A40:D40')
            cell = sheet2['A40']
            cell.value = 'merna nesigurnost pomnožena faktorom obuhvata k = 2, što za slučaj normalne raspodele greške'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet2.merge_cells('A41:D41')
            cell = sheet2['A41']
            cell.value = 'odgovara verovatnoći od približno 95 %.'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet2.merge_cells('A43:D43')
            cell = sheet2['A43']
            cell.value = '- Kraj Uverenja -'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(size=10)

            #********************************************************
            #Kraj jedne stranice
            #Ubacivanje podataka u stranicu 
            if(len(self.dataACI_M) != 0):
                a = len(self.dataACI)
                flag1 = True
                tmpACI = []
                tmpACI_M = []
                tmpACI_F = []
                cnt = 0
                for i in range(a):
                    tmpData = self.dataACI[i][0] #Vrednost
                    tmpData_M = self.dataACI_M[i][0] #Vrednost merena
                    tmpData_F = self.dataACI[i][1]
                    if(flag1 == True):
                        flag1 = False
                        sheet2.cell(row=6, column=3).value = self.dataACI[i][2]
                        sheet2.cell(row=6, column=1).value = self.dataACI[i][1]
                        sheet2.cell(row=7, column=1).value = self.dataACI[i][1]
                        cnt +=2
                    else:
                        sheet2.cell(row=6+cnt, column=3).value = self.dataACI[i][2]
                        sheet2.cell(row=6+cnt, column=1).value = self.dataACI[i][1]
                        cnt +=1

                    tmpACI_F.append(tmpData_F) 
                    r = len(tmpData)
                    for j in range(r):
                        tmpACI.append(tmpData[j])
                        tmpACI_M.append(tmpData_M[j])
                        
                    #sheet2.cell(row=6+i, column=2).value = tmpDCV[i]
            else:
                pass

            for i in range(len(tmpACI)):
                sheet2.cell(row=6+i, column=2).value = tmpACI[i]
                sheet2.cell(row=6+i, column=4).value = tmpACI_M[i]

                #Greska
                sheet2.cell(row=6+i, column=5).value = tmpACI_M[i] - tmpACI[i]

                sheet2.cell(row=6+i, column=1).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet2.cell(row=6+i, column=2).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet2.cell(row=6+i, column=3).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet2.cell(row=6+i, column=4).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet2.cell(row=6+i, column=5).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet2.cell(row=6+i, column=6).border = Border(top=None,bottom=None,right=thin,left=None)

                sheet2.cell(row=6+i, column=1).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet2.cell(row=6+i, column=2).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet2.cell(row=6+i, column=3).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet2.cell(row=6+i, column=4).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet2.cell(row=6+i, column=5).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet2.cell(row=6+i, column=6).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        else:
            pass

        #Otpor
        if(len(self.dataRES) != 0):
            sheet4 = book.create_sheet("Multimetar RES", self.cnt_sheet)
            self.cnt_sheet +=1
        
            sheet4['A3'].border = Border(top=thin,bottom=thin,right=thin,left=thin) 
            sheet4['B3'].border = Border(top=thin,bottom=thin,right=thin,left=thin) 
            sheet4['C3'].border = Border(top=thin,bottom=thin,right=None,left=thin) 
            sheet4['C3'].border = Border(top=thin,bottom=thin,right=None,left=None)
            sheet4['D3'].border = Border(top=thin,bottom=thin,right=thin,left=thin)
            sheet4['E3'].border = Border(top=thin,bottom=thin,right=None,left=thin) 
            sheet4['E3'].border = Border(top=thin,bottom=thin,right=None,left=None)
            sheet4['E3'].border = Border(top=thin,bottom=thin,right=None,left=None)  

            sheet4['A4'].border = Border(right=thin) 
            sheet4['B4'].border = Border(right=thin) 
            sheet4['C4'].border = Border(right=thin) 
            sheet4['D4'].border = Border(right=thin)
            sheet4['E4'].border = Border(right=thin)  

            sheet4['A5'].border = Border(bottom=double,right=thin) 
            sheet4['B5'].border = Border(bottom=double,right=thin) 
            sheet4['C5'].border = Border(bottom=double,right=thin) 
            sheet4['D5'].border = Border(bottom=double,right=thin)
            sheet4['E5'].border = Border(bottom=double,right=thin)  

            sheet4['A6'].border = Border(right=thin, top = None) 
            sheet4['B6'].border = Border(right=thin) 
            sheet4['C6'].border = Border(right=thin, top = None) 
            sheet4['D6'].border = Border(right=thin)
            sheet4['E6'].border = Border(right=thin) 

            
            sheet4['A7'].border = Border(right=thin) 
            sheet4['B7'].border = Border(right=thin) 
            sheet4['C7'].border = Border(right=thin) 
            sheet4['D7'].border = Border(right=thin)
            sheet4['E7'].border = Border(right=thin) 

            sheet4['A8'].border = Border(right=thin) 
            sheet4['B8'].border = Border(right=thin) 
            sheet4['C8'].border = Border(right=thin) 
            sheet4['D8'].border = Border(right=thin)
            sheet4['E8'].border = Border(right=thin) 

            #sirina celova
            sheet4.column_dimensions['A'].width = 20
            sheet4.column_dimensions['B'].width = 20
            sheet4.column_dimensions['C'].width = 20
            sheet4.column_dimensions['D'].width = 20
            sheet4.column_dimensions['E'].width = 20

            #Visina 
            sheet4.row_dimensions[4].height = 30
        
            
            cell = sheet4['A2']
            cell.value = 'Otpornost - Dvožično'


            cell = sheet4['A3']
            cell.value = 'Etalon'
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            
            sheet4.merge_cells('B3:C3')
            cell = sheet4['B3']
            cell.value = 'OE'
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            sheet4.merge_cells('D3:E3')
            cell = sheet4['D3']
            cell.value = 'Obrada rezultata'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet4['A4']
            cell.value = 'Pokazivanje etalona'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet4['B4']
            cell.value = 'Opseg instrumenta'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet4['C4']
            cell.value = 'Očitana vrednost sa instrumenta'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)


            cell = sheet4['D4']
            cell.value = 'G'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet4['E4']
            cell.value = 'U'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet4['A5']
            cell.value = 'Ω' #Jedinice merene velicine
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet4['B5']
            cell.value = 'Ω'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet4['C5']
            cell.value = 'Ω'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet4['D5']
            cell.value = 'Ω'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet4['B5']
            cell.value = 'Ω'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet4['E5']
            cell.value = 'Ω'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')


            cell = sheet4['A34']
            cell.value = 'OE'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet4['A35']
            cell.value = 'G'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet4['A36']
            cell.value = 'U'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet4['C34']
            cell.value = 'Objekt etaloniranja'

            cell = sheet4['C35']
            cell.value = 'Greška OE'

            cell = sheet4['C36']
            cell.value = 'Proširena (k=2) merna nesigurnost etaloniranja'


            sheet4.merge_cells('A39:D39')
            cell = sheet4['A39']
            cell.value = 'Merna nesigurnost iskazana u ovom Uverenju je proširena merna nesigurnost, gde je standardna'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet4.merge_cells('A40:D40')
            cell = sheet4['A40']
            cell.value = 'merna nesigurnost pomnožena faktorom obuhvata k = 2, što za slučaj normalne raspodele greške'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet4.merge_cells('A41:D41')
            cell = sheet4['A41']
            cell.value = 'odgovara verovatnoći od približno 95 %.'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet4.merge_cells('A43:D43')
            cell = sheet4['A43']
            cell.value = '- Kraj Uverenja -'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(size=10)

            #********************************************************
            #Kraj jedne stranice
            #Ubacivanje podataka u stranicu
            
            if(len(self.dataRES_M) != 0):
                a = len(self.dataRES)
                flag = True
                tmpRES = []
                tmpRES_M = []
                cnt = 0
                for i in range(a):
                    tmpData = self.dataRES[i][0]
                    tmpData_M = self.dataRES_M[i][0]
                    if(flag == True):
                        flag = False
                        sheet4.cell(row=6, column=2).value = self.dataRES[i][1]
                        cnt +=3
                    else:
                        sheet4.cell(row=6+cnt, column=2).value = self.dataRES[i][1]
                        cnt +=2 

                    r = len(tmpData)
                    for j in range(r):
                        tmpRES.append(tmpData[j])
                        tmpRES_M.append(tmpData_M[j])
            else:
                pass

            for i in range(len(tmpRES)):
                sheet4.cell(row=6+i, column=1).value = tmpRES[i]
                sheet4.cell(row=6+i, column=3).value = tmpRES_M[i]

                #Greska
                sheet4.cell(row=6+i, column=4).value = tmpRES_M[i] - tmpRES[i]

                sheet4.cell(row=6+i, column=1).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet4.cell(row=6+i, column=2).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet4.cell(row=6+i, column=3).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet4.cell(row=6+i, column=4).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet4.cell(row=6+i, column=5).border = Border(top=None,bottom=None,right=thin,left=None)

                sheet4.cell(row=6+i, column=1).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet4.cell(row=6+i, column=2).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet4.cell(row=6+i, column=3).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet4.cell(row=6+i, column=4).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet4.cell(row=6+i, column=5).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        else:
            pass  

        #Temperatura
        if(len(self.dataTEMP) != 0):

            sheet6 = book.create_sheet("Multimetar Temp", self.cnt_sheet)
            self.cnt_sheet +=1
           # ss_sheet6 = book['sheet6']
            #ss_sheet6.title = 'MULTIMETAR RES' #Ime sheet6a

            sheet6['A3'].border = Border(top=thin,bottom=thin,right=thin,left=thin) 
            sheet6['B3'].border = Border(top=thin,bottom=thin,right=thin,left=thin) 
            sheet6['C3'].border = Border(top=thin,bottom=thin,right=None,left=thin) 
            sheet6['C3'].border = Border(top=thin,bottom=thin,right=None,left=None)
            sheet6['D3'].border = Border(top=thin,bottom=thin,right=thin,left=thin)
            sheet6['E3'].border = Border(top=thin,bottom=thin,right=None,left=thin) 
            sheet6['E3'].border = Border(top=thin,bottom=thin,right=None,left=None)
            sheet6['E3'].border = Border(top=thin,bottom=thin,right=None,left=None)  

            sheet6['A4'].border = Border(right=thin) 
            sheet6['B4'].border = Border(right=thin) 
            sheet6['C4'].border = Border(right=thin) 
            sheet6['D4'].border = Border(right=thin)
            sheet6['E4'].border = Border(right=thin)  

            sheet6['A5'].border = Border(bottom=double,right=thin) 
            sheet6['B5'].border = Border(bottom=double,right=thin) 
            sheet6['C5'].border = Border(bottom=double,right=thin) 
            sheet6['D5'].border = Border(bottom=double,right=thin)
            sheet6['E5'].border = Border(bottom=double,right=thin)  

            sheet6['A6'].border = Border(right=thin, top = None) 
            sheet6['B6'].border = Border(right=thin) 
            sheet6['C6'].border = Border(right=thin, top = None) 
            sheet6['D6'].border = Border(right=thin)
            sheet6['E6'].border = Border(right=thin) 

            
            sheet6['A7'].border = Border(right=thin) 
            sheet6['B7'].border = Border(right=thin) 
            sheet6['C7'].border = Border(right=thin) 
            sheet6['D7'].border = Border(right=thin)
            sheet6['E7'].border = Border(right=thin) 

            sheet6['A8'].border = Border(right=thin) 
            sheet6['B8'].border = Border(right=thin) 
            sheet6['C8'].border = Border(right=thin) 
            sheet6['D8'].border = Border(right=thin)
            sheet6['E8'].border = Border(right=thin) 

            #sirina celova
            sheet6.column_dimensions['A'].width = 20
            sheet6.column_dimensions['B'].width = 20
            sheet6.column_dimensions['C'].width = 20
            sheet6.column_dimensions['D'].width = 20
            sheet6.column_dimensions['E'].width = 20

            #Visina 
            sheet6.row_dimensions[4].height = 30
        
            
            cell = sheet6['A2']
            tipSonde = self.dataTEMP[0]
            cell.value = 'Temperatura ' + tipSonde


            cell = sheet6['A3']
            cell.value = 'Etalon'
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            
            sheet6.merge_cells('B3:C3')
            cell = sheet6['B3']
            cell.value = 'OE'
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            sheet6.merge_cells('D3:E3')
            cell = sheet6['D3']
            cell.value = 'Obrada rezultata'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet6['A4']
            cell.value = 'Pokazivanje etalona'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet6['B4']
            cell.value = 'Opseg instrumenta'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet6['C4']
            cell.value = 'Očitana vrednost sa instrumenta'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)


            cell = sheet6['D4']
            cell.value = 'G'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet6['E4']
            cell.value = 'U'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet6['A5']
            cell.value = '°C' #Jedinice merene velicine
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet6['B5']
            cell.value = '°C'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet6['C5']
            cell.value = '°C'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet6['D5']
            cell.value = '°C'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet6['B5']
            cell.value = '°C'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet6['E5']
            cell.value = '°C'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')


            cell = sheet6['A34']
            cell.value = 'OE'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet6['A35']
            cell.value = 'G'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet6['A36']
            cell.value = 'U'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet6['C34']
            cell.value = 'Objekt etaloniranja'

            cell = sheet6['C35']
            cell.value = 'Greška OE'

            cell = sheet6['C36']
            cell.value = 'Proširena (k=2) merna nesigurnost etaloniranja'


            sheet6.merge_cells('A39:D39')
            cell = sheet6['A39']
            cell.value = 'Merna nesigurnost iskazana u ovom Uverenju je proširena merna nesigurnost, gde je standardna'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet6.merge_cells('A40:D40')
            cell = sheet6['A40']
            cell.value = 'merna nesigurnost pomnožena faktorom obuhvata k = 2, što za slučaj normalne raspodele greške'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet6.merge_cells('A41:D41')
            cell = sheet6['A41']
            cell.value = 'odgovara verovatnoći od približno 95 %.'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet6.merge_cells('A43:D43')
            cell = sheet6['A43']
            cell.value = '- Kraj Uverenja -'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(size=10)

            #********************************************************
            #Kraj jedne stranice
            #Ubacivanje podataka u stranicu
            
            if(len(self.dataTEMP_M) != 0):
                a = len(self.dataTEMP) - 1
                
                tmpData_M = self.dataTEMP_M
                for i in range(a):
                    #tmpData = self.dataTEMP[i]
                    
                    sheet6.cell(row=6+i, column=1).value = self.dataTEMP[i+1]
                    sheet6.cell(row=6+i, column=3).value = self.dataTEMP_M[i+1]

                    sheet6.cell(row=6+i, column=4).value = self.dataTEMP_M[i+1] - self.dataTEMP[i+1]

                    sheet6.cell(row=6+i, column=1).border = Border(top=None,bottom=None,right=thin,left=None)
                    sheet6.cell(row=6+i, column=2).border = Border(top=None,bottom=None,right=thin,left=None)
                    sheet6.cell(row=6+i, column=3).border = Border(top=None,bottom=None,right=thin,left=None)
                    sheet6.cell(row=6+i, column=4).border = Border(top=None,bottom=None,right=thin,left=None)
                    sheet6.cell(row=6+i, column=5).border = Border(top=None,bottom=None,right=thin,left=None)

                    sheet6.cell(row=6+i, column=1).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                    sheet6.cell(row=6+i, column=2).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                    sheet6.cell(row=6+i, column=3).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                    sheet6.cell(row=6+i, column=4).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                    sheet6.cell(row=6+i, column=5).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                
                #sheet6.cell(row=6+i-1, column=1).value = self.dataTEMP[i]
            else:
                pass    
        else:
            pass

        #Frekvencija
        if(len(self.dataFREQ) != 0):
            sheet5 = book.create_sheet("Multimetar Freq", self.cnt_sheet)
            self.cnt_sheet +=1
           
            sheet5['A3'].border = Border(top=thin,bottom=thin,right=thin,left=thin) 
            sheet5['B3'].border = Border(top=thin,bottom=thin,right=thin,left=thin) 
            sheet5['C3'].border = Border(top=thin,bottom=thin,right=None,left=thin) 
            sheet5['C3'].border = Border(top=thin,bottom=thin,right=None,left=None)
            sheet5['D3'].border = Border(top=thin,bottom=thin,right=thin,left=thin)
            sheet5['E3'].border = Border(top=thin,bottom=thin,right=None,left=thin) 
            sheet5['E3'].border = Border(top=thin,bottom=thin,right=None,left=None)
            sheet5['E3'].border = Border(top=thin,bottom=thin,right=None,left=None)  

            sheet5['A4'].border = Border(right=thin) 
            sheet5['B4'].border = Border(right=thin) 
            sheet5['C4'].border = Border(right=thin) 
            sheet5['D4'].border = Border(right=thin)
            sheet5['E4'].border = Border(right=thin)  

            sheet5['A5'].border = Border(bottom=double,right=thin) 
            sheet5['B5'].border = Border(bottom=double,right=thin) 
            sheet5['C5'].border = Border(bottom=double,right=thin) 
            sheet5['D5'].border = Border(bottom=double,right=thin)
            sheet5['E5'].border = Border(bottom=double,right=thin)  

            sheet5['A6'].border = Border(right=thin, top = None) 
            sheet5['B6'].border = Border(right=thin) 
            sheet5['C6'].border = Border(right=thin, top = None) 
            sheet5['D6'].border = Border(right=thin)
            sheet5['E6'].border = Border(right=thin) 

            
            sheet5['A7'].border = Border(right=thin) 
            sheet5['B7'].border = Border(right=thin) 
            sheet5['C7'].border = Border(right=thin) 
            sheet5['D7'].border = Border(right=thin)
            sheet5['E7'].border = Border(right=thin) 

            sheet5['A8'].border = Border(right=thin) 
            sheet5['B8'].border = Border(right=thin) 
            sheet5['C8'].border = Border(right=thin) 
            sheet5['D8'].border = Border(right=thin)
            sheet5['E8'].border = Border(right=thin) 

            #sirina celova
            sheet5.column_dimensions['A'].width = 20
            sheet5.column_dimensions['B'].width = 20
            sheet5.column_dimensions['C'].width = 20
            sheet5.column_dimensions['D'].width = 20
            sheet5.column_dimensions['E'].width = 20

            #Visina 
            sheet5.row_dimensions[4].height = 30
        
            
            cell = sheet5['A2']
            cell.value = 'Frekvencija'


            cell = sheet5['A3']
            cell.value = 'Etalon'
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            
            sheet5.merge_cells('B3:C3')
            cell = sheet5['B3']
            cell.value = 'OE'
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            sheet5.merge_cells('D3:E3')
            cell = sheet5['D3']
            cell.value = 'Obrada rezultata'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet5['A4']
            cell.value = 'Pokazivanje etalona'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet5['B4']
            cell.value = 'Opseg instrumenta'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            cell = sheet5['C4']
            cell.value = 'Očitana vrednost sa instrumenta'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)


            cell = sheet5['D4']
            cell.value = 'G'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet5['E4']
            cell.value = 'U'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet5['A5']
            cell.value = 'Hz' #Jedinice merene velicine
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet5['B5']
            cell.value = 'Hz'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet5['C5']
            cell.value = 'Hz'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet5['D5']
            cell.value = 'Hz'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet5['B5']
            cell.value = 'Hz'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet5['E5']
            cell.value = 'Hz'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')


            cell = sheet5['A34']
            cell.value = 'OE'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet5['A35']
            cell.value = 'G'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet5['A36']
            cell.value = 'U'
            cell.font = Font(italic=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

            cell = sheet5['C34']
            cell.value = 'Objekt etaloniranja'

            cell = sheet5['C35']
            cell.value = 'Greška OE'

            cell = sheet5['C36']
            cell.value = 'Proširena (k=2) merna nesigurnost etaloniranja'


            sheet5.merge_cells('A39:D39')
            cell = sheet5['A39']
            cell.value = 'Merna nesigurnost iskazana u ovom Uverenju je proširena merna nesigurnost, gde je standardna'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet5.merge_cells('A40:D40')
            cell = sheet5['A40']
            cell.value = 'merna nesigurnost pomnožena faktorom obuhvata k = 2, što za slučaj normalne raspodele greške'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet5.merge_cells('A41:D41')
            cell = sheet5['A41']
            cell.value = 'odgovara verovatnoći od približno 95 %.'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(italic=True, size=10)

            sheet5.merge_cells('A43:D43')
            cell = sheet5['A43']
            cell.value = '- Kraj Uverenja -'
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.font = Font(size=10)

            #********************************************************
            #Kraj jedne stranice
            #Ubacivanje podataka u stranicu
            
            if(len(self.dataFREQ_M) != 0):
                a = len(self.dataFREQ)
                flag = True
                tmpFREQ = []
                tmpFREQ_M = []
                cnt = 0
                for i in range(a):
                    tmpData = self.dataFREQ[i][0]
                    tmpData_M = self.dataFREQ_M[i][0]
                    if(flag == True):
                        flag = False
                        sheet5.cell(row=6, column=2).value = self.dataFREQ[i][1]
                        cnt +=1
                    else:
                        sheet5.cell(row=6+cnt, column=2).value = self.dataFREQ[i][1]
                        cnt +=1 

                    tmpFREQ.append(tmpData)
                    tmpFREQ_M.append(tmpData_M)
            else:
                pass

            for i in range(len(tmpFREQ)):
                sheet5.cell(row=6+i, column=1).value = tmpFREQ[i]
                sheet5.cell(row=6+i, column=3).value = tmpFREQ_M[i]

                #Greska
                sheet5.cell(row=6+i, column=4).value = tmpFREQ_M[i] - tmpFREQ[i]

                sheet5.cell(row=6+i, column=1).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet5.cell(row=6+i, column=2).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet5.cell(row=6+i, column=3).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet5.cell(row=6+i, column=4).border = Border(top=None,bottom=None,right=thin,left=None)
                sheet5.cell(row=6+i, column=5).border = Border(top=None,bottom=None,right=thin,left=None)

                sheet5.cell(row=6+i, column=1).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet5.cell(row=6+i, column=2).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet5.cell(row=6+i, column=3).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet5.cell(row=6+i, column=4).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet5.cell(row=6+i, column=5).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        else:
            pass

        self.dir_path = QFileDialog.getExistingDirectory(self,"Choose Directory","C:\\")
        now = datetime.now()
        s1 = now.strftime("%d.%m.%Y-%Hh %Mmin %Ssec")
        os.mkdir(self.dir_path+'/'+s1)  
        book.save(self.dir_path+'\\'+s1+'\\Multimetar.xlsx') #Cuvanje fajla
        self.ui.pushButtonSave.setEnabled(False)














        
        
            




        